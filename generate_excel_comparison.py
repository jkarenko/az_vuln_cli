#!/usr/bin/env python3
"""
Generate Excel workbook from vulnerability comparison data.
Creates formatted Excel sheets with filtering and sorting capabilities.
"""

import pandas as pd
import sys
import os
from pathlib import Path

def create_excel_comparison():
    """Create Excel workbook with all comparison data."""
    
    print("📊 Generating Excel comparison workbook...")
    
    # Check for available data files
    files_to_check = {
        'production_summary': 'reports/production/vulnerabilities_summary.csv',
        'dev_summary': 'reports/dev/vulnerabilities_summary.csv',
        'ACR_latest_summary': 'reports/latest/vulnerabilities_summary.csv', 
        'prod_vs_ACR': 'reports/comparison/detailed_vulnerability_comparison.csv',
        'dev_vs_ACR': 'reports/comparison/detailed_vulnerability_comparison_dev.csv',
        'production_os': 'reports/production/os_versions.csv',
        'dev_os': 'reports/dev/os_versions.csv',
        'latest_os': 'reports/latest/os_versions.csv'
    }
    
    available_files = {}
    for name, path in files_to_check.items():
        if Path(path).exists():
            available_files[name] = path
    
    if not available_files:
        print("❌ No vulnerability data found to export")
        return
    
    # Create Excel writer object
    output_file = 'reports/vulnerability_comparison_analysis.xlsx'
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    print(f"📁 Creating: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sheets_created = 0
        
        # Sheet 1: Production Summary
        if 'production_summary' in available_files:
            df = pd.read_csv(available_files['production_summary'])
            df.to_excel(writer, sheet_name='AKS Prod Summary', index=False)
            
            # Apply formatting
            worksheet = writer.sheets['AKS Prod Summary']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add filters to header row
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'AKS Prod Summary' sheet ({len(df)} rows)")
        
        # Sheet 2: Dev Summary
        if 'dev_summary' in available_files:
            df = pd.read_csv(available_files['dev_summary'])
            df.to_excel(writer, sheet_name='AKS Dev Summary', index=False)
            
            worksheet = writer.sheets['AKS Dev Summary']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'AKS Dev Summary' sheet ({len(df)} rows)")

        # Sheet 3: Latest Summary  
        if 'ACR_latest_summary' in available_files:
            df = pd.read_csv(available_files['ACR_latest_summary'])
            df.to_excel(writer, sheet_name='ACR latest Summary', index=False)
            
            worksheet = writer.sheets['ACR latest Summary']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'ACR latest Summary' sheet ({len(df)} rows)")
        
        # Sheet 4: Prod vs ACR latest Detailed
        if 'prod_vs_ACR' in available_files:
            df = pd.read_csv(available_files['prod_vs_ACR'])
            df.to_excel(writer, sheet_name='Prod vs ACR latest', index=False)
            
            worksheet = writer.sheets['Prod vs ACR latest']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'Prod vs ACR latest' sheet ({len(df)} rows)")
        
        # Sheet 5: Dev vs ACR latest Detailed
        if 'dev_vs_ACR' in available_files:
            df = pd.read_csv(available_files['dev_vs_ACR'])
            df.to_excel(writer, sheet_name='Dev vs ACR latest', index=False)
            
            worksheet = writer.sheets['Dev vs ACR latest']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'Dev vs ACR latest' sheet ({len(df)} rows)")
        
        # OS Versions sheets
        if 'production_os' in available_files:
            df = pd.read_csv(available_files['production_os'])
            df.to_excel(writer, sheet_name='Prod OS Versions', index=False)
            
            worksheet = writer.sheets['Prod OS Versions']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'Prod OS Versions' sheet ({len(df)} rows)")
            
        if 'dev_os' in available_files:
            df = pd.read_csv(available_files['dev_os'])
            df.to_excel(writer, sheet_name='Dev OS Versions', index=False)
            
            worksheet = writer.sheets['Dev OS Versions']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'Dev OS Versions' sheet ({len(df)} rows)")
            
        if 'latest_os' in available_files:
            df = pd.read_csv(available_files['latest_os'])
            df.to_excel(writer, sheet_name='Latest OS Versions', index=False)
            
            worksheet = writer.sheets['Latest OS Versions']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            worksheet.auto_filter.ref = worksheet.dimensions
            
            sheets_created += 1
            print(f"  ✅ Created 'Latest OS Versions' sheet ({len(df)} rows)")
        
        # Summary dashboard
        if sheets_created > 0:
            # Create a summary dashboard sheet
            summary_data = []
            
            if 'production_summary' in available_files:
                prod_df = pd.read_csv(available_files['production_summary'])
                high_count = len(prod_df[prod_df['Priority'] == 'High']) if 'Priority' in prod_df.columns else 0
                medium_count = len(prod_df[prod_df['Priority'] == 'Medium']) if 'Priority' in prod_df.columns else 0
                low_count = len(prod_df[prod_df['Priority'] == 'Low']) if 'Priority' in prod_df.columns else 0
                
                summary_data.append(['Production Environment', len(prod_df), high_count, medium_count, low_count])
            
            if 'latest_summary' in available_files:
                latest_df = pd.read_csv(available_files['latest_summary'])
                high_count = len(latest_df[latest_df['Priority'] == 'High']) if 'Priority' in latest_df.columns else 0
                medium_count = len(latest_df[latest_df['Priority'] == 'Medium']) if 'Priority' in latest_df.columns else 0
                low_count = len(latest_df[latest_df['Priority'] == 'Low']) if 'Priority' in latest_df.columns else 0
                
                summary_data.append(['Latest ACR Images', len(latest_df), high_count, medium_count, low_count])
            
            if 'dev_summary' in available_files:
                dev_df = pd.read_csv(available_files['dev_summary'])
                high_count = len(dev_df[dev_df['Priority'] == 'High']) if 'Priority' in dev_df.columns else 0
                medium_count = len(dev_df[dev_df['Priority'] == 'Medium']) if 'Priority' in dev_df.columns else 0
                low_count = len(dev_df[dev_df['Priority'] == 'Low']) if 'Priority' in dev_df.columns else 0
                
                summary_data.append(['Dev Environment', len(dev_df), high_count, medium_count, low_count])
            
            if summary_data:
                summary_df = pd.DataFrame(summary_data, columns=[
                    'Environment', 'Total Packages', 'High Priority', 'Medium Priority', 'Low Priority'
                ])
                summary_df.to_excel(writer, sheet_name='Summary Dashboard', index=False)
                
                worksheet = writer.sheets['Summary Dashboard']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                sheets_created += 1
                print(f"  ✅ Created 'Summary Dashboard' sheet")
        
        # Sheet 7: FAQ and help
        faq_data = [
            ['Question', 'Answer'],
            ['How do I read or use this data?', 'Each sheet contains vulnerability data for different environments. Use filters and sorting to focus on high-priority packages. The "Status" column in detailed sheets shows FIXED/NEW/IMPROVED/WORSENED/UNCHANGED to guide update decisions.'],
            ['How do I know what is fixed and where?', 'Look at the detailed comparison sheets (Prod vs ACR Latest, Dev vs ACR Latest). Packages with Status=FIXED will have their vulnerabilities eliminated by updating to :latest images. The "Fixed_Vulnerabilities" column lists the specific CVEs that will be resolved.'],
            ['How do I know what is fixed if I update an AKS image using the latest-tagged image from ACR?', 'Check the "Prod vs ACR Latest" or "Dev vs ACR Latest" sheet. Find your package and look at the Status column. If Status=FIXED, updating to :latest will eliminate all vulnerabilities for that package. If Status=IMPROVED, some vulnerabilities will be fixed (compare Vulnerability_Count_Prod vs Vulnerability_Count_Latest).'],
            ['What does each Status mean in the detailed sheets?', 'FIXED: Package vulnerabilities completely eliminated in :latest. NEW: Package has vulnerabilities only in :latest (new risk). IMPROVED: Fewer vulnerabilities in :latest than current. WORSENED: More vulnerabilities in :latest than current. UNCHANGED: Same number of vulnerabilities.'],
            ['Which packages should I prioritize for updates?', 'Focus on packages with Status=FIXED and High Priority first, as these give the biggest security improvement with no new risks. Then consider IMPROVED packages, weighing the reduction in vulnerabilities against any new ones introduced.'],
            ['How do I use the filters effectively?', 'Click the dropdown arrows in the header row. Filter by Priority=High to see critical issues first. Filter by Status=FIXED to see guaranteed improvements. Use multiple filters together (e.g., Priority=High AND Status=FIXED) for targeted analysis.'],
            ['What do the Priority levels mean?', 'High: Contains Critical or High severity CVEs - immediate attention needed. Medium: Contains Medium severity CVEs - should be addressed soon. Low: Contains only Low/Info severity CVEs - can be scheduled for routine maintenance.'],
            ['How current is this data?', 'Data reflects the state when the scan was last run. Vulnerability databases are updated daily, so re-run scans weekly or after major security announcements to get the latest data.'],
            ['What if a package shows as WORSENED?', 'This means the :latest version introduces more vulnerabilities than it fixes. Investigate the specific CVEs in the Fixed_Vulnerabilities column. Consider waiting for a newer version or applying targeted patches instead of updating.'],
            ['How do I plan my deployment strategy?', 'Start with packages that are FIXED and High Priority - these are no-regret updates. Group related packages together (e.g., all packages from the same base image). Test in dev environment first, then promote to production.'],
            ['What information is in each sheet?', 'Production/Dev/Latest Summary: Current vulnerabilities by environment. Detailed Comparison sheets: Package-by-package analysis of what changes when updating. OS Versions sheets: Operating system versions and EOSL status for each environment. Summary Dashboard: High-level overview across environments.'],
            ['What do the OS Versions sheets tell me?', 'These sheets show the operating system family, version, and End-of-Service-Life (EOSL) status for each container image. Images marked with EOSL=True are running on unsupported OS versions and should be updated urgently for security compliance.'],
            ['How do I identify images running on unsupported OS versions?', 'Check the OS Versions sheets and filter by EOSL=True. These images are running on operating systems that no longer receive security updates and represent significant security risks.'],
            ['How do I find specific packages or CVEs?', 'Use Ctrl/Cmd+F to search. You can also filter the Package or Fixed_Vulnerabilities columns. To find all instances of a CVE, search across the Fixed_Vulnerabilities column.'],
            ['What does "Affected_Images" tell me?', 'This shows which container images contain the vulnerable package. If multiple images contain the same package, updating the base image or package version will fix vulnerabilities across all affected images.'],
            ['How do I share this analysis with my team?', 'The Excel file is self-contained and can be shared directly. For presentations, copy key charts from the Summary Dashboard. For technical teams, share the relevant detailed comparison sheets with filters pre-applied.'],
            ['What if some expected packages are missing?', 'Missing packages might mean: 1) They have no vulnerabilities (good!), 2) They are not detected by Trivy scanning, or 3) The image was not included in the scan inventory. Check the original CSV files to confirm.']
        ]
        
        faq_df = pd.DataFrame(faq_data)
        faq_df.to_excel(writer, sheet_name='FAQ and Help', index=False, header=False)
        
        worksheet = writer.sheets['FAQ and Help']
        
        # Format FAQ sheet specially
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header row formatting
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col in range(1, 3):  # Columns A and B
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Question column formatting (bold)
        question_font = Font(bold=True, size=10)
        for row in range(2, len(faq_data) + 1):
            cell = worksheet.cell(row=row, column=1)
            cell.font = question_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Answer column formatting (wrapped text)
        answer_font = Font(size=10)
        for row in range(2, len(faq_data) + 1):
            cell = worksheet.cell(row=row, column=2)
            cell.font = answer_font
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set column widths for FAQ
        worksheet.column_dimensions['A'].width = 50  # Questions column
        worksheet.column_dimensions['B'].width = 80  # Answers column
        
        # Set row heights to accommodate wrapped text
        for row in range(2, len(faq_data) + 1):
            worksheet.row_dimensions[row].height = 60
            
        sheets_created += 1
        print(f"  ✅ Created 'FAQ and Help' sheet")
    
    print(f"\n✅ Excel workbook created successfully!")
    print(f"📁 Location: {output_file}")
    print(f"📊 Contains {sheets_created} worksheets with filtering enabled")
    print(f"\n💡 Open in Excel or LibreOffice for interactive analysis:")
    print(f"   - Use filters to focus on specific priorities")
    print(f"   - Sort columns to identify worst vulnerabilities")
    print(f"   - Compare sheets side-by-side for update decisions")

if __name__ == "__main__":
    try:
        create_excel_comparison()
    except ImportError:
        print("❌ Error: pandas and openpyxl are required")
        print("Install with: pip install pandas openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        sys.exit(1)